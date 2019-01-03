import datetime
import time
import os
import shutil
import openpyxl
import numpy as np
from binance.client import Client
from config import API_KEY, API_SECRET
from datetime import datetime
from config import time_dict
import matplotlib.pyplot as plt


EMAS_G = (3, 5, 8, 10, 12, 15)
EMAS_R = (30, 35, 40, 45, 50, 60)
DISPLACEMENT = 30
client = Client(API_KEY, API_SECRET)
cloud_calls = []


class Kline:
    def __init__(self, open_time, open_price, high_price, low_price, close_price, volume,
                 close_time, quote, no_trades, taker_base, taker_quote, ignore_val):
        self.low_price = float(low_price)
        self.close_price = float(close_price)
        self.volume = float(volume)
        self.close_time = close_time
        self.quote = float(quote)
        self.no_trades = int(no_trades)
        self.taker_base = float(taker_base)
        self.taker_quote = float(taker_quote)
        self.ignore_val = int(float(ignore_val))
        self.high_price = float(high_price)
        self.open_t = open_time
        self.open_price = float(open_price)

    def get_attributes(self):
        return time_from_ts(self.open_t), self.open_price, self.high_price, self.low_price, \
               self.close_price, self.volume, time_from_ts(self.close_time), self.quote, \
               self.no_trades, self.taker_base, self.taker_quote, self.ignore_val


def time_from_ts(timestamp, timespec='minutes'):
    st = datetime.fromtimestamp(timestamp / 1000)
    time_transformed = str(st.isoformat(timespec=timespec, sep=' '))
    return time_transformed


def get_coins_list():
    coins = [i['symbol'] for i in client.get_exchange_info()['symbols'] if 'BTC' in i['symbol']]
    return coins


def plot_ichimoku(klines, time, axs, market, conv_period=20, base_period=60, span2period=120, displacement=DISPLACEMENT,
                  plot_lines=True, plot_cloud=True):
    dist = time[1] - time[0]
    time_new = add_time(time, dist, displacement)

    try:
        axes = axs[0]
    except:
        axes = axs

    # defining ichimoku values
    try:
        conversion_line = ((get_period_highs(conv_period, klines)
                            + get_period_lows(conv_period, klines))
                           / 2)

        base_line = ((get_period_highs(base_period, klines)
                      + get_period_lows(base_period, klines))
                     / 2)
        senkou_a = (conversion_line[-len(base_line):] + base_line) / 2
        senkou_b = ((get_period_highs(span2period, klines)
                     + get_period_lows(span2period, klines))
                    / 2)
        chikou = get_chikou(displacement, klines)
        is_greater = senkou_a[-len(senkou_b):] >= senkou_b
        time_a = time_new[-len(senkou_b):]
    except ValueError as e:
        return print(f'The data provided is too short. Error occured: {e}')

    current_close_price = klines[-1].close_price
    current_open_price = np.mean([k.open_price for k in klines[-3:]])

    # plotting lines - optional
    if plot_lines:
        axes.plot(time[:len(chikou)], chikou, '-',
                    linewidth=1, markersize=1, color='g')
        axes.plot(time[-len(conversion_line):], conversion_line, '-',
                    linewidth=1, markersize=1, color='b')
        axes.plot(time[-len(base_line):], base_line, '-',
                    linewidth=1, markersize=1, color='r')
        axes.plot(time[:len(chikou)], chikou, '-',
                    linewidth=1, markersize=1, color='g')

    # plotting cloud
    if plot_cloud:
        axes.fill_between(time_a, senkou_a[-len(senkou_b):], senkou_b,
                            where=is_greater, color='g', alpha=0.1)
        axes.fill_between(time_a, senkou_a[-len(senkou_b):], senkou_b, where=~is_greater,
                            color='r', alpha=0.1)
        axes.plot(time_new[-len(senkou_a):], senkou_a, '-',
                    linewidth=0.5, markersize=1, color='g')
        axes.plot(time_new[-len(senkou_b):], senkou_b, '-',
                    linewidth=0.5, markersize=1, color='r')

    cross = cloud_cross(senkou_a, senkou_b, time_a, axs, displacement)
    cloud_close = is_close(senkou_a, senkou_b, current_open_price, current_close_price)
    thickness = cloud_thickness(current_open_price, current_close_price, senkou_a, senkou_b, DISPLACEMENT)

    # senkous crossing and checks if the nearby cloud is thin enough this checks future
    if cross and cloud_close:
        thin_cloud = is_thin(senkou_a, senkou_b, cross, current_close_price)
        if thin_cloud < 0.05:
            print(f'{market}\'s cloud cross is happening soon and the cloud is thin!')
            cloud_calls.append(market)

    # cloud thin and price close to the cloud - this checks current price
    if thickness < 0.025 or cloud_close:
        print(f'{market}\'s cloud is thin and/or the price is close to it!')
        cloud_calls.append(market)

    return senkou_a[-len(senkou_b):], senkou_b, abs(senkou_a[-len(senkou_b):] - senkou_b)


def plot_wicks(klines, axs, time_val):
    # function plots markets price diagram as bars(wicks)
    close_prices = [w.close_price for w in klines]
    a = np.array(close_prices[:-1])
    b = np.array(close_prices[1:])
    bool_filter = np.insert(a > b, 0, True)
    low_price, high_price, open_price, close_price = [], [], [], []

    [(low_price.append(w.low_price),
      high_price.append(w.high_price),
      close_price.append(w.close_price),
      open_price.append(w.open_price))
     for w in klines]

    time_np = np.array(time_val)
    low_price = np.array(low_price)
    high_price = np.array(high_price)
    open_price = np.array(open_price)
    close_price = np.array(close_price)

    try:
        axes = axs[0]
    except:
        axes = axs

    axes.vlines(time_np[bool_filter], low_price[bool_filter], high_price[bool_filter],
                  color='r', lw=1, zorder=5)
    axes.vlines(time_np[bool_filter], open_price[bool_filter], close_price[bool_filter],
                  color='r', lw=7, zorder=10)
    axes.vlines(time_np[~bool_filter], low_price[~bool_filter], high_price[~bool_filter],
                  color='g', lw=1, zorder=5)
    axes.vlines(time_np[~bool_filter], open_price[~bool_filter], close_price[~bool_filter],
                  color='g', lw=7, zorder=10)
    return bool_filter


def plot_vol(klines, axs, vol, bool_filter):
    # function plots volume on a figure below the main one
    max_val = np.max(vol)
    x, y = [], []
    [(x.append(w.open_t), y.append(w.volume / max_val)) for w in klines]
    axs[1].set_ylabel('volume', fontsize=14)
    axs[1].vlines(np.array(x)[bool_filter], 0, np.array(y)[bool_filter], color='r', lw=5)
    axs[1].vlines(np.array(x)[~bool_filter], 0, np.array(y)[~bool_filter], color='g', lw=5)
    axs[1].set_ylim([0, 1])


def cloud_cross(senkou_a, senkou_b, time_a, axs, displacement=DISPLACEMENT):
    # function checks if the senkou_a(green) and senkou_b(red)
    # are soon crossing each other (or recently did)
    i = -20
    if senkou_a[i] > senkou_b[i]:
        while senkou_a[i] > senkou_b[i]:
            i -= 1
            if i == -displacement - 12:
                i = None
                break
    else:
        while senkou_b[i] > senkou_a[i]:
            i -= 1
            if i == -displacement - 12:
                i = None
                break
    if i:
        cross_idx = i + 1
        axs[0].axvline(time_a[cross_idx])
        return cross_idx


def is_thin(senkou_a, senkou_b, point, close_price, steps=6):
    # func checks if the cloud is thin locally at defined point
    cloud_mean_r = np.mean(senkou_a[point:point + steps] - senkou_b[point:point + steps])
    cloud_mean_l = np.mean(senkou_a[point - steps:point] - senkou_b[point - steps:point])
    return min(cloud_mean_r / close_price, cloud_mean_l / close_price)


def rsi(klines, period=14):
    all_prices = [i.close_price for i in klines[:period+1]]
    all_prices_l = np.array(all_prices[:period])
    all_prices_r = np.array(all_prices[1:])

    d_price = np.array(all_prices_l) - np.array(all_prices_r)

    previous_avg_loss = np.sum(np.abs(d_price[all_prices_l > all_prices_r])) / period
    previous_avg_gain = np.sum(np.abs(d_price[all_prices_l < all_prices_r])) / period

    previous_price = klines[period+1].close_price
    avg_losses = [previous_avg_loss]
    avg_gains = [previous_avg_gain]
    for kline in klines[period+2:]:
        current_price = kline.close_price
        if previous_price > current_price:
            # loss
            gain = 0
            loss = previous_price - current_price
        else:
            # gain
            loss = 0
            gain = current_price - previous_price

        avg_gain = (previous_avg_gain * (period - 1) + gain) / period
        avg_gains.append(avg_gain)
        avg_loss = (previous_avg_loss * (period - 1) + loss) / period
        avg_losses.append(avg_loss)

        previous_avg_gain = avg_gain
        previous_avg_loss = avg_loss
        previous_price = current_price

    rs_val = np.array(avg_gains) / np.array(avg_losses)
    rsi_val = 100 - 100/(1+rs_val)
    return rsi_val


def plot_rsi(axs, klines, time_val):
    rsi_val = rsi(klines)
    axs[3].plot(time_val[-len(rsi_val):], rsi_val, '-',
                linewidth=1, markersize=1, color='b', alpha=0.5)
    axs[3].axhline(y=0, xmin=0, xmax=1, alpha=0.2)
    axs[3].set_ylabel('RSI', fontsize=14)
    return rsi_val


def cloud_thickness(open_price, close_price, senkou_a, senkou_b, disp, steps=3):
    # returns a minimum value of thickness as a % of coins current_price for further steps(wicks)
    # works good alone
    thicknesses = []
    price = (open_price + close_price) / 2
    for i in range(1, steps + 1):
        thk = abs(senkou_a[-disp + i] - senkou_b[-disp + i])
        thicknesses.append(thk)
    return min(thicknesses) / price


def all_clouds(klines, senkou_a, senkou_b, disp, steps=3):
    # returns cloud thickness at all points
    thicknesses = np.array([])
    for idx in range(-disp+5, -len(klines)+disp, -1):
        price = (klines[idx].open_price + klines[idx].close_price) / 2
        area_utc = np.sum(np.abs(senkou_a[-idx-steps: -idx + steps] - senkou_b[-idx-steps: -idx + steps]))
        thicknesses = np.insert(thicknesses, 0, area_utc / price)
    return thicknesses


def is_close(senkou_a, senkou_b, open_price, close_price, point=DISPLACEMENT):
    # checks if the current price is close to the cloud
    counter = 0.02
    limiter = max(senkou_a[-point], senkou_b[-point])
    price = (open_price + close_price) / 2
    if open_price >= limiter:
        if (abs(open_price - senkou_a[-point]) < counter * price) | (
                abs(open_price - senkou_b[-point]) < counter * price):
            return True
        else:
            return False
    else:
        if (abs(close_price - senkou_a[-point]) < counter * price) | (
                abs(close_price - senkou_b[-point]) < counter * price):
            return True
        else:
            return False


def plot_macd(axs, klines, time_val):
    signal, ema_9 = macd(klines)
    max_val = np.max(np.append(signal, ema_9))
    signal_t = signal / max_val
    ema_9_t = ema_9 / max_val
    axs[2].plot(time_val[-len(signal_t)+20:], signal_t[20:], '-',
                linewidth=1, markersize=1, color='b', alpha=0.5)
    axs[2].plot(time_val[-len(ema_9_t)+20:], ema_9_t[20:], '-',
                linewidth=1, markersize=1, color='r', alpha=0.5)
    axs[2].axhline(y=0, xmin=0, xmax=1, alpha=0.2)
    axs[2].set_ylabel('MACD', fontsize=14)

    return signal[-len(ema_9):] / ema_9


def plot_gmma(axs, klines, time_val):
    # function plots gmma-indicator based on given emas periods
    emas_dict = {}
    for period in EMAS_G:
        color = 'g'
        try:
            ema_val = ema(klines, period)
        except IndexError as e:
            return print(e)
        axs[0].plot(time_val[period:], ema_val, '-',
                    linewidth=1, markersize=1, color=color, alpha=0.5)
        emas_dict[period] = ema_val
    for period in EMAS_R:
        color = 'r'

        try:
            ema_val = ema(klines, period)
        except IndexError as e:
            return print(e)

        axs[0].plot(time_val[period:], ema_val, '-',
                    linewidth=1, markersize=1, color=color, alpha=0.5)
        emas_dict[period] = ema_val
    return emas_dict


def gmma_cross(emas, price, periods=6):
    # function checks if the short and long emas are crossing each other
    counter = 0.005
    shorts = []
    longs = []
    for i in range(periods):
        gen_g = (emas[k][-i] for k in emas.keys() if k in EMAS_G)
        gen_r = (emas[k][-i] for k in emas.keys() if k in EMAS_R)
        short = np.mean([i for i in gen_g])
        long = np.mean([i for i in gen_r])
        shorts.append(np.mean(short))
        longs.append(np.mean(long))
    return np.any(np.abs(np.array(shorts) - np.array(longs)) < counter * price)


def gmma_tightness(klines):
    all_emas = EMAS_G + EMAS_R
    long_dict = {}
    short_dict = {}

    for ema_val in all_emas:
        current_ema = ema(klines, ema_val)
        if ema_val in EMAS_G:
            short_dict[ema_val] = current_ema
        if ema_val in EMAS_R:
            long_dict[ema_val] = current_ema

    length_short = np.min(np.array([len(i) for i in list(short_dict.values())]))
    cut_vals_sh = [i[-length_short:] for i in list(short_dict.values())]
    sorted_vals_sh = np.stack(tuple(cut_vals_sh), axis=-1)
    max_vals_sh = np.max(sorted_vals_sh, axis=1)
    min_vals_sh = np.min(sorted_vals_sh, axis=1)
    short_tightness = abs(max_vals_sh - min_vals_sh)

    length_long = np.min(np.array([len(i) for i in list(long_dict.values())]))
    cut_vals_long = [i[-length_long:] for i in list(long_dict.values())]
    sorted_vals_long = np.stack(tuple(cut_vals_long), axis=-1)
    max_vals_long = np.max(sorted_vals_long, axis=1)
    min_vals_long = np.min(sorted_vals_long, axis=1)
    long_tightness = abs(max_vals_long - min_vals_long)

    width = abs(max_vals_sh[-len(min_vals_long)] - min_vals_long)
    return short_tightness[-len(long_tightness):], long_tightness, width


def gmma_compression(emas, price, periods=6, ema_type=EMAS_R):
    # checks the differences between prices of long/short emas
    counter = 0.008
    final = []
    try:
        for i in range(periods):
            result = []
            gen = (emas[k][-i] for k in emas.keys() if k in ema_type)
            values = [i for i in gen]
            for j in range(len(values) - 1):
                diff = np.abs(values[j] - values[j + 1])
                result.append(diff)
            final.append(np.sum(result))
    except AttributeError as e:
        print(e)
        return False
    return np.any(np.array(final) < counter * price)


def sma(klines, period):
    # func calculates and returns a sma list-object
    period_sma = []
    for i in range(period - 1, len(klines)):
        values_sma = []
        for k in klines[i - period + 1:i]:
            values_sma.append(k.close_price)
        period_sma.append(np.mean(values_sma))
    return period_sma


def ema(klines, period):
    # func calculates and returns an ema list-object
    sma_val = sma(klines, period)[0]
    multiplier = (2 / (period + 1))
    ema_values = []
    first_val = klines[period]
    ema_previous = (first_val.close_price - sma_val) * multiplier + sma_val
    ema_values.append(ema_previous)
    for i in range(period + 1, len(klines)):
        ema_current = (klines[i].close_price - ema_previous) * multiplier + ema_previous
        ema_values.append(ema_current)
        ema_previous = ema_current
    return np.array(ema_values)


def macd(klines, per1=12, per2=26, per3=9):
    ema_12 = np.array(ema(klines, per1))
    ema_26 = np.array(ema(klines, per2))
    sub_res = ema_12[-len(ema_26):] - ema_26

    values_sma = []
    for k in klines[:per3]:
        values_sma.append(k.close_price)
    sma_val = np.mean(values_sma)

    ema_values = []
    first_val = sub_res[per3]
    multiplier = (2 / (per3 + 1))
    ema_previous = (first_val - sma_val) * multiplier + sma_val
    ema_values.append(ema_previous)
    for i in range(per3 + 1, len(sub_res)):
        ema_current = (sub_res[i] - ema_previous) * multiplier + ema_previous
        ema_values.append(ema_current)
        ema_previous = ema_current
    return sub_res, ema_values


def organize_calls(calls, dirname=time.strftime("%Y-%m-%d %H_%M"), add=None):
    # function moves called coins to seperate dirnames
    # todo - only create the dir if actual calls exist
    if calls:
        path = os.getcwd()
        # current_time = time.strftime("%Y-%m-%d %H_%M")
        source = os.path.join(path, 'plots')
        os.chdir(source)
        files = os.listdir(os.getcwd())
        if not add:
            dir_name = dirname
        else:
            dir_name = dirname + f' {add}'

        for coin in calls:
            # directory = f'{folder}'
            if not os.path.exists(dir_name):
                os.makedirs(dir_name)
                print(f'{dir_name} directory created.')
            for f in files:
                if f.startswith(f'{coin}'):
                    try:
                        shutil.move(f'{f}', os.path.join(dir_name, f))
                        print(f'File {f} moved to {dir_name} directory.')
                    except FileNotFoundError as e:
                        print(e)
                        continue
        os.chdir(path)


# todo - could recreate the function to match both highs/lows and other stuff
def get_period_highs(period, klines):
    # function returns the lowest price value at given (past) period
    period_high = []
    for i in range(period, len(klines)):
        values_high = []
        # todo- implement recursion here
        for j in klines[i - period:i]:
            values_high.append(j.high_price)
        period_high.append(np.max(values_high))
    return np.array(period_high)


def get_period_lows(period, klines):
    # function returns the lowest price value at given (past) period
    period_low = []
    for i in range(period, len(klines)):
        values_low = []
        # todo- implement recursion here
        for j in klines[i - period:i]:
            values_low.append(j.low_price)
        period_low.append(np.min(values_low))
    return np.array(period_low)


def get_chikou(period, klines):
    # function returns chikou values (closing prices at given period)
    values_chi = [i.close_price for i in klines[period:]]
    return np.array(values_chi)


def get_time(klines):
    # function returns time values for given data
    time_val = [i.open_t for i in klines]
    return np.array(time_val)


def add_time(time_list, time_amount, repeater):
    # function adds time at given DISPLACEMENT - necessary for proper broadcasting
    for _ in range(repeater):
        time_list = np.append(time_list, time_list[-1] + time_amount)
    return time_list


def close_ema():
    # check if open_price is a little more than conv_line (more than 0.2% and less than 0.4% of open price
    pass


def baboo_indicator():
    # while True
    # todo-making new checks
    # def buy_walls_emerging
    # def buy_walls_dropping
    # def sell_walls_emerging
    # def sell_walls dropping
    pass


def vol_check():
    # whats the 24h/4h/1h vol increase/decrease
    pass


def update_excel(filename, coin_list):
    # creating/opening excel file containing current prices
    # the excel file will later be helpful at checking if the calls are working
    if os.path.isfile(filename):
        print('file exists')
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    else:
        print('File doesnt exist. Creating new one called coins.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Coins_table'
        ws['A1'] = 'Date/Coins'
        wb.save(filename)

    # data input to excel
    last_row = ws.max_row
    last_col = ws.max_column

    ws[f'A{last_row+1}'] = datetime.now()
    coins = {key.value: value for value, key in enumerate(list(ws.rows)[0], 2)}
    i = 1
    for coin in coin_list:
        bin_data = list(client.get_historical_klines_generator(symbol=f'{coin}',
                                                               interval=getattr(Client, time_dict['4H']),
                                                               start_str='8 hours ago UTC'))[-1]
        data = Kline(*bin_data)
        if coin not in coins.keys():
            print(f'{coin} doesnt exist yet in a table - adding new column...')
            ws.cell(row=1, column=last_col + i).value = coin
            ws.cell(row=last_row + 1, column=last_col + i).value = (data.open_price + data.close_price) / 2
            i += 1
        else:
            ws.cell(row=last_row + 1, column=coins[coin] - 1).value = (data.open_price + data.close_price) / 2

    wb.save(filename)
    wb.close()
    return print(f"Excel file {filename} created/updated.")
